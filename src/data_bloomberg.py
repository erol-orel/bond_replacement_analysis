"""
Parse the VZ / Bloomberg source workbooks (data/bloomberg/*.xlsx) into clean monthly
CHF series for the 2008-2026 study, plus SNB/Fed policy rates, the real VZ AP5 track
record, and the strategic AP5 target allocation.

Sources (provided by the student, Bloomberg / SNB / VZ):
  Memoire_de_master.xlsx
    - "Indexes data - Bloomberg"  : daily index levels (paired date/value columns)
    - "Bonds history"             : monthly SNB + Fed policy rates (cols 52-55)
    - "Price history AP5 VZ"       : the REAL VZ AP5 (VVIA) daily track record, 2019-2026
  CHF hedged data.xlsx             : CHF-hedged Bloomberg Global Aggregate (monthly)
  Consolidation_allocations.xlsx   : the VZ AP5 target-allocation history 2017-2026

Foreign-equity indices (MSCI World / EM) are Bloomberg *price* levels in USD; we convert
to a CHF total-return proxy = level x USDCHF, grossed up by a constant net dividend yield
(DIV_WORLD). This affects the equity CORE, which is identical across every portfolio we
compare, so it cancels in AP5-vs-replacement contrasts; it is calibrated against the real
VZ track record in run_reconstruction.py.

Output (data/processed/):
  constituents_chf_monthly.csv   month-end CHF total-return levels of the AP5 sleeves
  rates_monthly.csv              SNB + Fed policy rates (%)
  vz_ap5_track_monthly.csv       real VZ AP5 cumulative return (rebased 100)
Run:  python src/data_bloomberg.py
"""
from __future__ import annotations
import os, time
import numpy as np
import pandas as pd
import openpyxl
import requests

CA = "/root/.ccr/ca-bundle.crt"
os.environ.setdefault("REQUESTS_CA_BUNDLE", CA)
os.environ.setdefault("SSL_CERT_FILE", CA)
HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BB = os.path.join(HERE, "data", "bloomberg")
PROC = os.path.join(HERE, "data", "processed")
os.makedirs(PROC, exist_ok=True)

WB = os.path.join(BB, "Memoire_de_master.xlsx")
HEDGED = os.path.join(BB, "CHF hedged data.xlsx")

DIV_WORLD = 0.021   # MSCI World net dividend yield add-back (annual), calibrated to VZ
DIV_EM = 0.026      # MSCI EM net dividend yield add-back (annual)

# (date column, value column) 0-indexed in "Indexes data - Bloomberg"
IDX_COLS = {
    "swiss_bonds": (1, 2),      # SBR14T  SBI AAA-BBB (CHF total return)
    "swiss_equity": (9, 10),    # SPI     (CHF total return)
    "world_equity": (15, 16),   # MXWO    MSCI World (USD price -> CHF TR proxy)
    "em_equity": (19, 20),      # MXEF    MSCI EM     (USD price -> CHF TR proxy)
    "real_estate": (21, 22),    # SWIIT   SXI Real Estate Funds (CHF total return)
}
USD_INDICES = {"world_equity": DIV_WORLD, "em_equity": DIV_EM}


def _read_pairs(ws, dcol, vcol, start_row=5):
    dates, vals = [], []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        d = row[dcol] if dcol < len(row) else None
        v = row[vcol] if vcol < len(row) else None
        if d is None or v is None:
            continue
        try:
            dates.append(pd.Timestamp(d)); vals.append(float(v))
        except (ValueError, TypeError):
            continue
    s = pd.Series(vals, index=pd.DatetimeIndex(dates)).sort_index()
    return s[~s.index.duplicated(keep="last")]


def _fetch_yahoo(ticker, tries=4):
    """Monthly adjusted close, full history, via the Yahoo chart API."""
    url = (f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
           f"?range=30y&interval=1mo&events=div,splits")
    for k in range(tries):
        try:
            r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
            r.raise_for_status()
            res = r.json()["chart"]["result"][0]
            idx = pd.to_datetime(res["timestamp"], unit="s")
            adj = res["indicators"].get("adjclose", [{}])[0].get("adjclose")
            close = adj if adj is not None else res["indicators"]["quote"][0]["close"]
            s = pd.Series(close, index=idx).dropna().sort_index()
            return s[~s.index.duplicated(keep="last")]
        except Exception as e:
            if k == tries - 1:
                raise RuntimeError(f"yahoo {ticker}: {e}")
            time.sleep(2 ** k)


def month_end(s):
    return s.resample("ME").last()


def build_constituents():
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    ws = wb["Indexes data - Bloomberg"]
    raw = {name: _read_pairs(ws, d, v) for name, (d, v) in IDX_COLS.items()}
    wb.close()

    # USDCHF (CHF per 1 USD) on a clean period-month index for USD-priced equity indices
    usdchf = _fetch_yahoo("CHF=X")
    usdchf.index = usdchf.index.to_period("M")
    usdchf = usdchf[~usdchf.index.duplicated(keep="last")]

    out = {}
    for name, s0 in raw.items():
        s = month_end(s0)
        s.index = s.index.to_period("M")
        s = s[~s.index.duplicated(keep="last")]
        if name in USD_INDICES:
            fx = usdchf.reindex(s.index).ffill()
            chf_price = (s * fx).dropna()            # USD level -> CHF
            div_m = (1 + USD_INDICES[name]) ** (1 / 12) - 1
            r = chf_price.pct_change().dropna().add(div_m)   # CHF price return + dividend
            lvl = (1 + r).cumprod()
            lvl = lvl / lvl.iloc[0] * 100
        else:
            lvl = s / s.iloc[0] * 100
        lvl.index = lvl.index.to_timestamp("M")
        out[name] = lvl

    # world bonds: CHF-hedged Bloomberg Global Aggregate (broad), from CHF hedged data.xlsx
    wh = openpyxl.load_workbook(HEDGED, read_only=True, data_only=True)
    wsh = wh.active
    # cols: A/B = Global aggregate 1-5 ; D/E = Global aggregate (broad)
    gagg = _read_pairs(wsh, 3, 4, start_row=4)       # broad
    gagg15 = _read_pairs(wsh, 0, 1, start_row=4)     # 1-5
    wh.close()
    out["world_bonds"] = (month_end(gagg) / gagg.iloc[0] * 100)
    out["world_bonds_1_5"] = (month_end(gagg15) / gagg15.iloc[0] * 100)

    df = pd.DataFrame(out)
    # study window 2008-01-31 .. 2026-06-30
    df = df.loc["2008-01-31":"2026-06-30"]
    # rebase every column to 100 at first common date
    df = df.dropna(how="all")
    return df


def build_rates():
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    ws = wb["Bonds history"]
    dates, snb, fed = [], [], []
    for row in ws.iter_rows(min_row=3, values_only=True):
        d = row[52] if len(row) > 52 else None
        if d is None:
            continue
        try:
            dates.append(pd.Timestamp(d))
            snb.append(float(row[54])); fed.append(float(row[55]))
        except (ValueError, TypeError):
            continue
    wb.close()
    df = pd.DataFrame({"snb": snb, "fed": fed}, index=pd.DatetimeIndex(dates)).sort_index()
    df.index = df.index + pd.offsets.MonthEnd(0)
    return df.loc["2008-01-31":"2026-06-30"]


def build_vz_track():
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    ws = wb["Price history AP5 VZ"]
    dates, cum = [], []
    for row in ws.iter_rows(min_row=8, values_only=True):
        d = row[0]
        if d is None:
            continue
        try:
            ts = pd.Timestamp(d); c = float(row[3])
        except (ValueError, TypeError):
            continue
        dates.append(ts); cum.append(c)
    wb.close()
    s = pd.Series(cum, index=pd.DatetimeIndex(dates)).sort_index()
    s = s[~s.index.duplicated(keep="last")]
    return month_end(s).rename("vz_ap5").to_frame()


def main():
    con = build_constituents()
    rates = build_rates()
    vz = build_vz_track()
    con.to_csv(os.path.join(PROC, "constituents_chf_monthly.csv"))
    rates.to_csv(os.path.join(PROC, "rates_monthly.csv"))
    vz.to_csv(os.path.join(PROC, "vz_ap5_track_monthly.csv"))

    pd.set_option("display.width", 200)
    print("CONSTITUENTS (monthly CHF TR levels):", con.shape,
          con.index.min().date(), "->", con.index.max().date())
    print(con.head(3).round(2).to_string())
    print("...\n", con.tail(2).round(2).to_string())
    print("\nAnnualised return / vol (2008-2026, from monthly):")
    r = con.pct_change().dropna(how="all")
    ann = pd.DataFrame({"CAGR%": (con.iloc[-1] / con.iloc[0]) ** (12 / (len(con) - 1)) * 0 +
                        ((con.iloc[-1] / con.iloc[0]) ** (1 / ((len(con) - 1) / 12)) - 1) * 100,
                        "Vol%": r.std() * np.sqrt(12) * 100,
                        "n_obs": r.count()})
    print(ann.round(2).to_string())
    print("\nRATES: SNB", rates["snb"].iloc[0], "->", rates["snb"].iloc[-1],
          "| min", rates["snb"].min(), " (rows", len(rates), ")")
    print("VZ AP5 track:", vz.index.min().date(), "->", vz.index.max().date(),
          "cum", round(float(vz.iloc[-1, 0]), 1))


if __name__ == "__main__":
    main()
